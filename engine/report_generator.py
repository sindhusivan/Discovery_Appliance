"""
TAO Discovery Platform — Excel Report Generator
Produces a branded multi-sheet workbook with uniform borders.
"""

import pandas as pd
import sqlite3
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime

DB_PATH = "/opt/discovery-appliance/data/discovery.db"
REPORTS_DIR = "/opt/discovery-appliance/reports"

def generate_enterprise_excel():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'{REPORTS_DIR}/TAO_Infrastructure_Report_{timestamp}.xlsx'

    conn = sqlite3.connect(DB_PATH)

    df_vms = pd.read_sql_query("""
        SELECT vm_name, power_state, configured_os, running_os, guest_hostname,
               vcpu, ram_gb, num_disks, disk_sizes_gb, provisioned_space_gb,
               used_space_gb, ip_address, mac_address, portgroup, host_name,
               cluster_name, vcenter_tags, vmware_tools_status, vmware_tools_version,
               scan_timestamp
        FROM virtual_machines ORDER BY vm_name
    """, conn)

    df_hosts = pd.read_sql_query("""
        SELECT hostname, cluster_name, ip_address, vendor, model, cpu_model,
               cpu_sockets, cpu_cores_per_socket, total_cpu_cores, ram_gb,
               esxi_version, esxi_build, connection_state, power_state,
               num_nics, num_hbas, scan_timestamp
        FROM esxi_hosts ORDER BY hostname
    """, conn)

    df_storage = pd.read_sql_query("""
        SELECT name, capacity_gb, free_gb, datastore_type, accessible, scan_timestamp
        FROM datastores ORDER BY name
    """, conn)

    df_deps = pd.read_sql_query("""
        SELECT src_vm_name, src_ip, dst_vm_name, dst_ip,
               dst_port, protocol, flow_count, first_seen, last_seen
        FROM dependencies ORDER BY flow_count DESC
    """, conn)

    conn.close()

    # Insert clean sequential ID column
    for df in [df_vms, df_hosts, df_storage, df_deps]:
        df.insert(0, 'ID', range(1, len(df) + 1))

    # ── STYLES ──
    hdr_fill   = PatternFill(start_color="0D2461", end_color="0D2461", fill_type="solid")
    hdr_font   = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    hdr_border = Border(
        left=Side(style="thin",   color="1155CC"),
        right=Side(style="thin",  color="1155CC"),
        top=Side(style="thin",    color="1155CC"),
        bottom=Side(style="medium", color="00D4FF")
    )
    alt_fill   = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    thin_side  = Side(style="thin", color="E2E8F0")
    thin_bdr   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    id_fill    = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    id_font    = Font(bold=True, name="Calibri", size=10, color="1155CC")
    norm_font  = Font(name="Calibri", size=10)
    ctr        = Alignment(horizontal="center", vertical="center")
    lft        = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    sheets = {
        "Virtual Machines":  df_vms,
        "ESXi Hosts":        df_hosts,
        "Datastores":        df_storage,
        "App Dependencies":  df_deps,
    }

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.freeze_panes = "A2"

            # Auto-fit column widths
            for col in ws.columns:
                col_letter = col[0].column_letter
                max_len = max(
                    (len(str(cell.value)) if cell.value is not None else 0)
                    for cell in col
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            # Style header row
            for cell in ws[1]:
                cell.fill   = hdr_fill
                cell.font   = hdr_font
                cell.border = hdr_border
                cell.alignment = ctr

            # Style data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = alt_fill if row_idx % 2 == 0 else None
                for cell in row:
                    cell.border = thin_bdr
                    cell.font   = norm_font
                    cell.alignment = lft
                    if fill:
                        cell.fill = fill

            # Style ID column
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
                for cell in row:
                    cell.fill      = id_fill
                    cell.font      = id_font
                    cell.alignment = ctr
                    cell.border    = thin_bdr

    return output_file


# Alias for backward compatibility
generate_all_reports = generate_enterprise_excel
